from driver import Driver

from openpyxl import Workbook


def save_data(skins_data: list, items_data: list) -> None:
    wb = Workbook()

    for skin_data in skins_data:
        wb.create_sheet(skin_data['name'][0:31])
        ws = wb[skin_data['name'][0:31]]

        ws['A1'] = 'price'
        ws['B1'] = 'float'
        ws['C1'] = 'pattern'
        ws['D1'] = 'sticker'
        ws['E1'] = 'link'

        ws['A2'] = skin_data['price']
        ws['B2'] = skin_data['float']
        ws['C2'] = skin_data['pattern']
        ws['D2'] = skin_data['stickers']
        ws['E2'] = skin_data['url']

    row = 2

    wb.create_sheet('all_items')
    ws = wb['all_items']

    ws['A1'] = 'Item'
    ws['B1'] = 'price'
    ws['C1'] = 'float'
    ws['D1'] = 'pattern'
    ws['E1'] = 'stickers'
    ws['F1'] = 'link'

    for item_data in items_data:
        try:
            ws['A' + str(row)] = item_data['name']
            ws['B' + str(row)] = item_data['price']
            ws['C' + str(row)] = item_data['float']
            ws['D' + str(row)] = item_data['pattern']
            ws['E' + str(row)] = item_data['stickers']
            ws['F' + str(row)] = item_data['url']
            row += 1
        except:
            pass

    wb.save('data.xlsx')


def get_link(file_name: str) -> list:
    with open(file_name, 'r') as file:
        file_data = file.read()

    return ['https://market.csgo' + link for link in ''.join(file_data.split('\n')).split('https://market.csgo')][1:]


def main():
    driver = Driver()

    while True:
        username = input('Введите свой логин: ')
        password = input('Введите свой пароль: ')

        if driver.log_in(username, password):
            break
        else:
            print('Вы не правильно вели логин или пароль')

    print('Вход был успешным')

    driver.parsing_skins(get_link('data.txt'))

    driver.parsing_items()
    save_data(driver.skins_data, driver.items_data)


if __name__ == '__main__':
    main()
