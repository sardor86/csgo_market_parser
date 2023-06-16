from openpyxl import Workbook


class Data:
    def __init__(self, path: str):
        self.file_name = path
        self.wb = Workbook()

    def save_skin(self, skins_list: list) -> None:
        for skin_data in skins_list:
            self.wb.create_sheet(skin_data['name'][0:31])
            ws = self.wb[skin_data['name'][0:31]]

            ws['A1'] = 'price'
            ws['B1'] = 'float'
            ws['C1'] = 'pattern'
            ws['D1'] = 'link'

            ws['A2'] = skin_data['price']
            ws['B2'] = skin_data['float']
            ws['C2'] = skin_data['pattern']
            ws['D2'] = skin_data['url']

            row = 3

            for item_data in skin_data['items_data']:
                ws['A' + str(row)] = item_data['price']
                ws['B' + str(row)] = item_data['float']
                ws['C' + str(row)] = item_data['pattern']
                ws['D' + str(row)] = item_data['url']

    def save_items(self, items_list: list) -> None:
        row = 2

        self.wb.create_sheet('all_items')
        ws = self.wb['all_items']

        ws['A1'] = 'Item'
        ws['B1'] = 'price'
        ws['C1'] = 'float'
        ws['D1'] = 'pattern'
        ws['E1'] = 'link'

        for item_data in items_list:
            ws['A' + str(row)] = item_data['name']
            ws['B' + str(row)] = item_data['price']
            ws['C' + str(row)] = item_data['float']
            ws['D' + str(row)] = item_data['pattern']
            ws['E' + str(row)] = item_data['url']
            row += 1

    def save(self):
        self.wb.save(self.file_name)

    def __del__(self):
        self.save()
